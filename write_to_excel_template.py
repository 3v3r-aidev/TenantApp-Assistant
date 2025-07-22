import openpyxl
import re
import traceback
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from io import BytesIO
from datetime import datetime, date
import pandas as pd
from openpyxl.styles import Alignment
from extract_tenant_data import normalize_all_dates, normalize_date_string

def calc_age(dob_str: str) -> str | int:
    if not dob_str:
        return ""
    
    try:
        dob_str = str(dob_str).strip()
    except:
        return "Invalid DOB"

    # Check if it's an Excel serial date (pure digits, float-like)
    if re.match(r"^\d+(\.0+)?$", dob_str):
        try:
            dob = datetime(1899, 12, 30) + pd.to_timedelta(float(dob_str), unit="D")
            dob = dob.date()
            today = date.today()
            return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        except:
            return "Invalid DOB"

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            dob = datetime.strptime(dob_str, fmt).date()
            today = date.today()
            return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        except ValueError:
            continue
    return "Invalid DOB"


def lookup_property_info(address: str, reference_file="PropertyInfo.xlsx"):
    try:
        if not address:
            return None, None

        wb_info = openpyxl.load_workbook(reference_file, data_only=True)
        ws_info = wb_info.active

        address_prefix = " ".join(address.strip().lower().split()[:3])

        for row in ws_info.iter_rows(min_row=2):
            c_val = str(row[2].value).strip().lower() if row[2].value else ""
            c_prefix = " ".join(c_val.split()[:3])
            if address_prefix == c_prefix:
                p_number = row[1].value  # Column B
                sqft = row[3].value      # Column D
                return p_number, sqft
        return None, None
    except Exception as e:
        print("❌ Error in lookup_property_info:", e)
        return None, None

# ───────────────────────────────────────────────────────────────────────────────
# 1. write_flattened_to_template  (adds strict input-type guard)
# ───────────────────────────────────────────────────────────────────────────────
def write_flattened_to_template(
    data,
    template_path="templates/Tenant_Template.xlsx",
    summary_header=None,
):
    try:
        data = normalize_all_dates(data)
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Property Address
        property_address = str(data.get("Property Address", "") or "").strip()
        if ws.oddHeader.left is None:
            ws.oddHeader.left = openpyxl.worksheet.header_footer.HeaderFooterItem()
        ws.oddHeader.left.text = property_address

        if summary_header:
            if ws.oddHeader.center is None:
                ws.oddHeader.center = openpyxl.worksheet.header_footer.HeaderFooterItem()
            existing = ws.oddHeader.center.text or ""
            lines = existing.split("\n")
            new_line = f"Date={summary_header}"
            if len(lines) >= 3:
                lines[2] = new_line
            else:
                lines += [""] * (2 - len(lines)) + [new_line]
            ws.oddHeader.center.text = "\n".join(lines)

        ws["E3"] = property_address
        ws["E4"] = data.get("Move-in Date", "")
        ws["E5"] = str(data.get("Monthly Rent", "")).replace("$", "").strip()

        # PropertyInfo lookup
        try:
            prop_df = pd.read_excel("PropertyInfo.xlsx", header=None, dtype=str)
            addr_prefix = " ".join(property_address.strip().lower().split()[:3])
            mask = prop_df[2].fillna("").str.lower().apply(lambda x: " ".join(x.split()[:3])) == addr_prefix
            match = prop_df[mask]
            if not match.empty:
                ws["G3"] = match.iloc[0, 1]
                ws["G7"] = match.iloc[0, 3]
        except Exception as e:
            print(f"Warning: Failed lookup – {e}")

        # Safely write fields
        try:
            ws["F10"] = data.get("Rep Name", "")
            ws["J9"] = data.get("Rep Phone", "")
            ws["J10"] = data.get("Rep Email", "")
            ws["F14"] = data.get("FullName", "")
            ws["F15"] = data.get("Email", "")
            ws["F16"] = data.get("PhoneNumber", "")
            ws["F17"] = data.get("SSN", "")
            ws["F18"] = data.get("DriverLicenseNumber", "")
            ws["F19"] = data.get("DOB", "")
            ws["F20"] = calc_age(data.get("DOB", ""))
            
            num_occupants = str(data.get("No of Occupants", ""))
            ws["F21"] = str(num_occupants)  # Always write to F21

            # Only write to G21 if there's a second applicant (i.e., if G14 is filled)
            second_applicant = str(ws["G14"].value or "").strip()
            if second_applicant:
               ws["G21"] = num_occupants

            ws["F22"] = data.get("No of Children", "")
            ws["F23"] = data.get("Applicant's Current Address", "")
            ws["F24"] = data.get("Landlord or Property Manager's Name", "")
            ws["F25"] = data.get("Landlord Phone", "")
            ws["F27"] = data.get("Applicant's Current Employer", "")
            ws["F28"] = data.get("Employer Address", "")
            ws["F29"] = f"{data.get('Employment Verification Contact', '')} {data.get('Employer Phone', '')}".strip()
            ws["F30"] = data.get("Start Date", "")
            ws["F31"] = data.get("Gross Monthly Income", "")
            ws["F32"] = data.get("Position", "")
        except Exception as e:
            print(f"❌ Field assignment error: {e}")

        # Vehicles
        try:
            v_types = str(data.get("Vehicle Type", "") or "").split(",")
            v_makes = str(data.get("Vehicle Make", "") or "").split(",")
            v_models = str(data.get("Vehicle Model", "") or "").split(",")
            v_years = str(data.get("Vehicle Year", "") or "").split(",")
            vehicle_lines = [
                f"{t.strip()} {m.strip()} {mo.strip()} {y.strip()}".strip()
                for t, m, mo, y in zip(v_types, v_makes, v_models, v_years)
                if any([t.strip(), m.strip(), mo.strip(), y.strip()])
            ]
            ws["F34"] = "\n".join(vehicle_lines)
            ws["F34"].alignment = Alignment(wrap_text=True)
        except Exception as e:
            print(f"⚠️ Vehicle info error: {e}")

        # Vehicle Payments
        try:
            v_payments = str(data.get("Vehicle Monthly Payment", "")).split(",")
            cleaned = [p.replace("$", "").replace(",", "").strip() for p in v_payments]
            numeric = [float(p) for p in cleaned if p.replace(".", "", 1).isdigit()]
            ws["F35"] = sum(numeric) if len(numeric) > 1 else (numeric[0] if numeric else "")
        except Exception as e:
            print(f"⚠️ Vehicle payment error: {e}")

        # Ratios
        try:
            rent = float(str(data.get("Monthly Rent", "0")).replace("$", "").replace(",", "").strip() or 0)
            gross = float(str(data.get("Gross Monthly Income", "0")).replace("$", "").replace(",", "").strip() or 0)
            co_total = 0
            for app in data.get("Co-applicants", []):
                if isinstance(app, dict):
                    val = str(app.get("Gross Monthly Income", "")).replace("$", "").replace(",", "").strip()
                    try:
                        co_total += float(val)
                    except:
                        pass
            net_total = gross + co_total
            ws["J3"] = f"{gross / rent:.2f}" if rent > 0 else ""
            ws["J4"] = f"{net_total / rent:.2f}" if rent > 0 else ""
        except Exception as e:
            print(f"⚠️ Ratio calculation error: {e}")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        def generate_filename(address):
            cleaned = re.sub(r"[^\w\s]", "", str(address))
            words = cleaned.strip().split()
            word_part = "_".join(words[1:3]) if len(words) >= 3 else "_".join(words[:2]) if len(words) >= 2 else "tenant"
            return f"{word_part}_{datetime.now():%Y%m%d}_app.xlsx"

        return output, generate_filename(property_address)

    except Exception:
        print("❌ Error in write_flattened_to_template:")
        traceback.print_exc()
        return None, None
  
# ───────────────────────────────────────────────────────────────────────────────
# 2. write_multiple_applicants_to_template  (adds per-row type guard)
# ───────────────────────────────────────────────────────────────────────────────
def write_multiple_applicants_to_template(
    df,
    template_path="templates/Tenant_Template_Multiple.xlsx",
    summary_header=None,
):
    import traceback
    try:
        first_row = normalize_all_dates(df.iloc[0].to_dict())
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # ── Property Info ─────────────────────────────
        property_address = str(first_row.get("Property Address", "") or "").strip()

        # Initialize left header if missing
        if ws.oddHeader.left is None:
            ws.oddHeader.left = openpyxl.worksheet.header_footer.HeaderFooterItem()
        ws.oddHeader.left.text = property_address

        # Initialize center header if summary_header is used
        if summary_header:
            if ws.oddHeader.center is None:
                ws.oddHeader.center = openpyxl.worksheet.header_footer.HeaderFooterItem()
            existing = ws.oddHeader.center.text or ""
            lines = existing.split("\n")
            new_line = f"Date={summary_header}"
            if len(lines) >= 3:
                lines[2] = new_line
            else:
                lines += [""] * (2 - len(lines)) + [new_line]
            ws.oddHeader.center.text = "\n".join(lines)

        ws["E3"] = property_address
        ws["E4"] = first_row.get("Move-in Date", "")
        ws["E5"] = str(first_row.get("Monthly Rent", "")).replace("$", "").strip()
        ws["F10"] = first_row.get("Rep Name", "")
        ws["J9"] = first_row.get("Rep Phone", "")
        ws["J10"] = first_row.get("Rep Email", "")

        # ── PropertyInfo.xlsx Lookup ─────────────────────────────
        try:
            prop_df = pd.read_excel("PropertyInfo.xlsx", header=None, dtype=str)
            addr_prefix = " ".join(property_address.strip().lower().split()[:3])
            mask = prop_df[2].fillna("").str.lower().apply(lambda x: " ".join(x.split()[:3])) == addr_prefix
            match = prop_df[mask]
            if not match.empty:
                ws["G3"] = match.iloc[0, 1]
                ws["G7"] = match.iloc[0, 3]
        except Exception as e:
            print(f"Warning: Failed PropertyInfo lookup – {e}")

        # ── Gross & Net Ratio Calculation ──────────────────────────────
        rent_val = first_row.get("Monthly Rent", "")
        gross_val = first_row.get("Gross Monthly Income", "")

        rent_str = str(rent_val).replace("$", "").replace(",", "").strip() if rent_val is not None else ""
        gross_str = str(gross_val).replace("$", "").replace(",", "").strip() if gross_val is not None else ""

        try:
            rent = float(rent_str) if rent_str else 0
        except Exception as e:
            print(f"⚠️ Invalid rent value '{rent_str}': {e}")
            rent = 0
        try:
            gross = float(gross_str) if gross_str else 0
        except Exception as e:
            print(f"⚠️ Invalid gross income value '{gross_str}': {e}")
            gross = 0

        try:
            gross_ratio = f"{gross / rent:.2f}" if rent > 0 else ""
        except Exception as e:
            print(f"⚠️ Gross ratio calculation failed: {e}")
            gross_ratio = ""

        # Co-applicant Total Income for net ratio
        co_total = 0
        try:
            for _, row_series in df.iterrows():
                row = row_series.to_dict()
                val = str(row.get("Gross Monthly Income", "")).replace("$", "").replace(",", "").strip()
                try:
                    if val and float(val) != gross:  # avoid double-counting main applicant
                        co_total += float(val)
                except:
                    continue
        except Exception as e:
            print(f"⚠️ Failed to calculate co-applicant income: {e}")

        try:
            net_total = gross + co_total
            net_ratio = f"{net_total / rent:.2f}" if rent > 0 else ""
        except Exception as e:
            print(f"⚠️ Net ratio calculation failed: {e}")
            net_ratio = ""

        ws["J3"] = gross_ratio
        ws["J4"] = net_ratio

        # ── Fill applicant columns ──────────────────────────────────────
        col_starts = ["F", "I", "L", "O", "R", "U", "X", "AA", "AD", "AG"]
        start_row = 14

        for idx, (_, row_series) in enumerate(df.iterrows()):
            if idx >= len(col_starts):
                break

            if not hasattr(row_series, "to_dict"):
                raise TypeError(f"Row {idx} must be Series, got {type(row_series)}")

            row = normalize_all_dates(row_series.to_dict())
            col = col_starts[idx]

            def write(offset, value):
                try:
                    ws[f"{col}{start_row + offset}"] = value or ""
                except Exception as e:
                    print(f"⚠️ Failed to write value '{value}' at {col}{start_row + offset}: {e}")

            write(0, row.get("FullName"))
            write(1, row.get("Email"))
            write(2, row.get("PhoneNumber"))
            write(3, row.get("SSN"))
            write(4, row.get("DriverLicenseNumber"))
            write(5, row.get("DOB"))
            try:
                write(6, calc_age(row.get("DOB", "")))
            except Exception as e:
                print(f"⚠️ Failed to calculate age: {e}")
                write(6, "")
            write(7, str(row.get("No of Occupants", "")))
            write(8, row.get("No of Children", ""))
            write(9, row.get("Applicant's Current Address"))
            write(10, row.get("Landlord or Property Manager's Name"))
            write(11, row.get("Landlord Phone"))
            write(13, row.get("Applicant's Current Employer"))
            write(14, row.get("Employer Address"))
            write(15, f"{row.get('Employment Verification Contact', '')} {row.get('Employer Phone', '')}".strip())
            write(16, row.get("Start Date"))
            write(17, row.get("Gross Monthly Income"))
            write(19, row.get("Position"))

            try:
                v_types = str(row.get("Vehicle Type", "") or "").split(",")
                v_makes = str(row.get("Vehicle Make", "") or "").split(",")
                v_models = str(row.get("Vehicle Model", "") or "").split(",")
                v_years = str(row.get("Vehicle Year", "") or "").split(",")

                vehicle_lines = [
                    f"{t.strip()} {m.strip()} {mo.strip()} {y.strip()}".strip()
                    for t, m, mo, y in zip(v_types, v_makes, v_models, v_years)
                    if any([t.strip(), m.strip(), mo.strip(), y.strip()])
                ]

                vehicle_cell = f"{col}{start_row + 20}"
                ws[vehicle_cell] = "\n".join(vehicle_lines) if vehicle_lines else ""
                ws[vehicle_cell].alignment = openpyxl.styles.Alignment(wrap_text=True)
            except Exception as e:
                print(f"⚠️ Error building vehicle info: {e}")

            try:
                v_payments = str(row.get("Vehicle Monthly Payment", "")).split(",")
                cleaned_vals = [p.replace("$", "").replace(",", "").strip() for p in v_payments if p.strip()]
                numeric_vals = [float(p) for p in cleaned_vals if p.replace(".", "", 1).isdigit()]
                total_payment = sum(numeric_vals) if len(numeric_vals) > 1 else (numeric_vals[0] if numeric_vals else "")
                write(21, total_payment)
            except Exception as e:
                print(f"⚠️ Error calculating vehicle payment: {e}")
                write(21, "")

        output = BytesIO()
        try:
            wb.save(output)
            output.seek(0)
        except Exception as e:
            print(f"❌ Failed to save workbook: {e}")
            return None, None

        def generate_filename(address):
            try:
                cleaned = re.sub(r"[^\w\s]", "", str(address))
                words = cleaned.strip().split()
                word_part = (
                    "_".join(words[1:3]) if len(words) >= 3 else
                    "_".join(words[:2]) if len(words) >= 2 else
                    "tenant"
                )
                return f"{word_part}_{datetime.now():%Y%m%d}_app.xlsx"
            except Exception as e:
                print(f"⚠️ Filename generation failed: {e}")
                return f"tenant_{datetime.now():%Y%m%d}_app.xlsx"

        return output, generate_filename(property_address)

    except Exception:
        print("❌ Error in write_multiple_applicants_to_template:")
        traceback.print_exc()
        return None, None

# ───────────────────────────────────────────────────────────────────────────────
# 3. write_to_summary_template  (now type-safe)
# ───────────────────────────────────────────────────────────────────────────────
def write_to_summary_template(
    flat_data,
    output_path,
    summary_template_path="templates/App_Summary_Template.xlsx",
) -> None:
    """
    Writes one applicant’s key facts into App_Summary_Template.xlsx.
    """
    try:
        if isinstance(flat_data, dict):
            pass
        elif hasattr(flat_data, "to_dict"):
            flat_data = flat_data.to_dict()
        else:
            raise TypeError(f"write_to_summary_template expected dict/Series, got {type(flat_data)}")

        flat_data = normalize_all_dates(flat_data)

        try:
            wb = load_workbook(summary_template_path)
            ws = wb.active
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {e}")

        # ── Meta Sheet and Counter ─────────────────────────────
        try:
            meta_ws = wb["_Meta"] if "_Meta" in wb.sheetnames else wb.create_sheet("_Meta")
            if "_Meta" not in wb.sheetnames:
                wb.move_sheet(meta_ws, offset=len(wb.sheetnames))
                meta_ws.sheet_state = "hidden"
                meta_ws["A1"] = "counter"

            TEST_MODE = True
            TEST_START_VALUE = 636
            counter = TEST_START_VALUE if TEST_MODE else (meta_ws["B1"].value or TEST_START_VALUE) + 1
            meta_ws["B1"] = counter
            ws["B1"] = datetime.now().strftime(f"APP-{counter}-%Y-%m-%d-%H%M%S")
        except Exception as e:
            print(f"⚠️ Failed to update _Meta counter: {e}")

        # ── Safe Rent/Income Parsing ─────────────────────────
        try:
            rent_val = flat_data.get("Monthly Rent", "")
            rent_str = str(rent_val).replace("$", "").replace(",", "").strip() if rent_val else ""
            rent = float(rent_str) if rent_str else 0
        except Exception as e:
            print(f"⚠️ Invalid rent value: {e}")
            rent = 0

        try:
            gross_val = flat_data.get("Gross Monthly Income", "")
            gross_str = str(gross_val).replace("$", "").replace(",", "").strip() if gross_val else ""
            gross = float(gross_str) if gross_str else 0
        except Exception as e:
            print(f"⚠️ Invalid gross income value: {e}")
            gross = 0

        # ── Gross & Net Ratio Calculation (Updated) ──────────────
        try:
            co_total = 0
            for app in flat_data.get("Co-applicants", []):
                if isinstance(app, dict):
                    val = str(app.get("Gross Monthly Income", "")).replace("$", "").replace(",", "").strip()
                    try:
                        if val:
                            co_total += float(val)
                    except:
                        continue
            net_total = gross + co_total
            gross_ratio = f"{gross / rent:.2f}" if rent > 0 else ""
            net_ratio = f"{net_total / rent:.2f}" if rent > 0 else ""
        except Exception as e:
            print(f"⚠️ Failed to compute income ratios: {e}")
            gross_ratio = ""
            net_ratio = ""

        # ── Occupant Count Correction ─────────────────────────
        try:
            co_applicants = flat_data.get("Co-applicants", [])
            occupants = flat_data.get("E. Occupant Information", [])

            co_applicant_count = sum(
                1 for c in co_applicants 
                if isinstance(c, dict) and (c.get("Name") or c.get("FullName"))
            )
            occupant_count = sum(
                1 for o in occupants 
                if isinstance(o, dict) and (o.get("Name") or o.get("FullName"))
            )

            total_occupants = 1 + co_applicant_count + occupant_count
        except Exception as e:
            print(f"⚠️ Failed to compute total occupants: {e}")
            total_occupants = flat_data.get("No of Occupants", "")

        # ── Vehicle String Assembly ───────────────────────────
        vehicle = ""
        try:
            vehicle_lines = []
            if isinstance(flat_data.get("F. Vehicle Information:"), list):
                for v in flat_data["F. Vehicle Information:"]:
                    if not isinstance(v, dict):
                        continue
                    line = " ".join(
                        str(v.get(k, "")).strip()
                        for k in ("Type", "Year", "Make", "Model")
                        if v.get(k)
                    ).strip()
                    if line:
                        vehicle_lines.append(line)
            else:
                v_types = str(flat_data.get("Vehicle Type", "")).split(",")
                v_years = str(flat_data.get("Vehicle Year", "")).split(",")
                v_makes = str(flat_data.get("Vehicle Make", "")).split(",")
                v_models = str(flat_data.get("Vehicle Model", "")).split(",")
                for t, y, mke, mdl in zip(v_types, v_years, v_makes, v_models):
                    line = f"{t.strip()} {y.strip()} {mke.strip()} {mdl.strip()}"
                    if line:
                        vehicle_lines.append(line)
            vehicle = "\n".join(vehicle_lines)
        except Exception as e:
            print(f"⚠️ Error assembling vehicle info: {e}")

        # ── Animal String Assembly ───────────────────────────
        animals = ""
        try:
            animal_lines = []
            if isinstance(flat_data.get("G. Animals"), list):
                for a in flat_data["G. Animals"]:
                    if not isinstance(a, dict):
                        continue
                    line = " | ".join(
                        f"{label}: {a.get(key)}" for label, key in [
                            ("Type & Breed", "Type and Breed"),
                            ("Name", "Name"),
                            ("Color", "Color"),
                            ("Weight", "Weight"),
                            ("Age", "Age in Yrs"),
                            ("Gender", "Gender"),
                        ] if a.get(key)
                    )
                    if line:
                        animal_lines.append(line)
            else:
                default_animals = flat_data.get("Animal Details", flat_data.get("No of Animals", ""))
                if isinstance(default_animals, str) and default_animals.strip():
                    animal_lines.append(default_animals.strip())
            animals = "\n".join(animal_lines)
        except Exception as e:
            print(f"⚠️ Error assembling animal info: {e}")

        # ── Map to Summary Fields ────────────────────────────
        try:
            write_map = {
                "B2": flat_data.get("Property Address", ""),
                "B3": flat_data.get("Monthly Rent", ""),
                "B4": flat_data.get("Move-in Date", ""),
                "B5": flat_data.get("Application Fee", ""),
                "B6": f"{gross_ratio}/{net_ratio}",
                "B7": str(total_occupants),
                "B8": flat_data.get("Rent", ""),
                "B9": flat_data.get("Applicant's Current Employer", ""),
                "B12": vehicle,
                "B13": animals,
            }

            for cell, value in write_map.items():
                ws[cell] = value

        except Exception as e:
            print(f"❌ Error writing fields to worksheet: {e}")
            traceback.print_exc()

        try:
            wb.save(output_path)
        except Exception as e:
            raise RuntimeError(f"❌ Failed to save summary workbook: {e}")

    except Exception as final_error:
        print("❌ write_to_summary_template failed:")
        traceback.print_exc()
        raise final_error






