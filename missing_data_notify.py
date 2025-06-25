from openpyxl import load_workbook
from email.message import EmailMessage
import smtplib
import os
from dotenv import load_dotenv

# Load credentials from .env
load_dotenv()
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")

# Define the required fields and their row positions (row 4 is actual Excel row 4)
required_fields = {
    "Name": 4,
    "Email": 5,
    "Cell Phone": 6,
    "SSN": 7,
    "DL No.": 8,
    "DOB": 9,
    "Current Address": 13,
    "Employer": 17,
    "Gross Monthly Income": 21
}

# Get the correct column letters for each applicant index
def get_column_letters(index):
    base = 6 + index * 3  # F=6, I=9, L=12, etc.
    return chr(64 + base), chr(64 + base + 1)

# Safely read from a cell
def read_cell(ws, col, row):
    return str(ws[f"{col}{row}"].value or "").strip()

# Email notification logic
def send_email(to_email, missing_fields):
    msg = EmailMessage()
    msg['Subject'] = "Missing Information in Your Rental Application"
    msg['From'] = EMAIL_USER
    msg['To'] = to_email

    msg.set_content(
        f"""Dear Applicant,

We reviewed your rental application and noticed the following missing information:

{', '.join(missing_fields)}

Please provide the missing details at your earliest convenience so we can continue processing your application.

Thank you,
Property Management Team
        """
    )

    with smtplib.SMTP("smtp.ionos.com", 587) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_USER, EMAIL_PASS)
        smtp.send_message(msg)
        print(f"\U0001F4E7 Email sent to: {to_email}")

# Main checker function
def check_excel_and_notify(template_file):
    wb = load_workbook(template_file)
    ws = wb.active

    applicant_idx = 0
    while True:
        col1, _ = get_column_letters(applicant_idx)
        email = read_cell(ws, col1, 14)

        if not email:
            break  # Stop loop if no email found

        missing = []
        for field, row in required_fields.items():
            value = read_cell(ws, col1, row)
            value_clean = value.strip().lower()
            if not value_clean or value_clean in {"n/a", "-", "none", "null"}:
                missing.append(field)

        if missing:
            print(f"⚠️ Missing fields for {email}: {missing}")
            send_email(email, missing)
        else:
            print(f"✅ Applicant {applicant_idx + 1} has all required fields.")

        applicant_idx += 1

# Example usage
if __name__ == "__main__":
    check_excel_and_notify("Tenant_Template_Filled.xlsx")